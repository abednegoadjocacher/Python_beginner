import openpyxl as xl# type:ignore
from openpyxl.chart import BarChart, Reference # type:ignore
from pathlib import Path


workbook = xl.load_workbook('simplefile.xlsx')
sheet = workbook['Sheet1']
cell = sheet['a1']
#_cell = sheet.cell(1 ,1)# Another way to access cell.

for row in range(2 , sheet.max_row +1):
    cell = sheet.cell(row ,3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    #print(cell.value)
    #print(row)

_values = Reference(sheet, min_row=2,max_row=sheet.max_row, min_col=4, max_col=4)
chart= BarChart()
chart.add_data(_values)
sheet.add_chart(chart, 'f5')
workbook.save('simple.xlsx')


#print(cell)
#print(_cell)
print(sheet.max_row)

# path = Path()
# #print(path.glob("*.py")) # the generates objects of all files
# for file in path.glob('*.py'): # To search for all file with "py" extension
#     print(file)